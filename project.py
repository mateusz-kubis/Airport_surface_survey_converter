import os
import openpyxl
from pathlib import Path
from reportlab.lib.pagesizes import letter, inch
from reportlab.platypus import Image, SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib import colors
import tkinter as tk
from tkinter.filedialog import askopenfilename
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont   
from reportlab.lib.styles import ParagraphStyle

initial_file = Path().absolute()

assets_catalog = f'{Path().absolute()}/assets'

size_of_cell = 0.6*inch

data = []
images = {}
title = []

def code_to_pictogram(code):
  for files in os.walk(assets_catalog):
      if f'{code}.png' in files[2]:
          return 1
      else:
          return 0

def open_file():
    file_path = askopenfilename(filetypes=[("Excel 2007-365 (macros)", "*.xlsm"),("Excel 2007-365", "*.xlsx"),("Excel 97-2003", "*.xls")])
    if not file_path:
        print("No file...")
        return
    else:
        print(file_path)
        initial_file = file_path

    max_cols = 0
    max_rows = 0

    try:
        for file in os.listdir(assets_catalog):
            if file.endswith(".png"):
                images[Path(file).stem] = Image(f'{assets_catalog}/{file}')
                images[Path(file).stem].drawWidth = size_of_cell
                images[Path(file).stem].drawHeight = size_of_cell

        wb = openpyxl.load_workbook(initial_file)
        ws = wb.active
        
        firstRowData = ['','']
        for i in range(1,max_cols):
            firstRowData.append(chr(64+i))
        data.append(firstRowData)

        col_num = 0
        row_data = []

        for row in ws.iter_rows(min_row=2, min_col=1, max_col=1, max_row=ws.max_row):
            max_rows = row[0].row
            if ((ws.cell(row=row[0].row, column=1).value) is None):
                max_rows -= 1
                break

        for column in ws.iter_cols(min_row=1, min_col=2, max_row=1, max_col=ws.max_column):
            max_cols = column[0].column
            if ((ws.cell(row=1, column=column[0].column).value) is None):
                max_cols -= 1
                break

        for row in ws.iter_rows(min_row=1, min_col=1, max_row=max_rows, max_col=max_cols):
            if col_num != 0:
                row_data = [col_num]
            for cell in row:
                if(cell.value is None):
                    row_data.append('')
                else:
                    if (code_to_pictogram(cell.value) == 1):
                        row_data.append([images[cell.value]])
                    else:
                        row_data.append(cell.value)
            data.append(row_data)
            col_num += 1            
        wb.close()

    except Exception as e:
        print(f"Błąd: {e}")
        input("Wciśnij Enter, aby zamknąć...")

def save_file():
    try:
        elements = []
        pdfmetrics.registerFont(TTFont('DejaVuSerif', 'DejaVuSerif.ttf'))

        style = ParagraphStyle(
            name='Normal',
            fontName='DejaVuSerif',
            fontSize=20,
            alignment=1,
            leading=50
        )

        doc = SimpleDocTemplate("out.pdf", pagesize=letter, title="Stan techniczny nawierzchni")
        tab = Table(data, size_of_cell, size_of_cell)
        tab.setStyle(TableStyle([
            ('ALIGN', (0,0), (-1,0), 'CENTER'),   # first row
            ('VALIGN', (0,0), (-1,0), 'MIDDLE'),
            ('TEXTCOLOR', (0,0), (-1,0), colors.green),

            ('ALIGN', (0,1), (0,-1), 'CENTER'),   # first column
            ('VALIGN', (0,1), (0,-1), 'MIDDLE'),
            ('TEXTCOLOR', (0,1), (0,-1), colors.green),

            ('ALIGN', (1,1), (-1,1), 'RIGHT'),   # second row
            ('VALIGN', (1,1), (-1,1), 'BOTTOM'),

            ('ALIGN', (1,2), (1,-1), 'RIGHT'),   # second column
            ('VALIGN', (1,2), (1,-1), 'BOTTOM'),

            ('ALIGN', (2,2), (-1,-1), 'CENTER'),
            ('VALIGN', (2,2), (-1,-1), 'MIDDLE'),
            ('INNERGRID', (2,2), (-1,-1), 0.25, colors.black),
            ('BOX', (2,2), (-1,-1), 0.25, colors.black)
            ]))

        tab._argW[0] = 0.5 * size_of_cell
        tab._argW[1] = 0.5 * size_of_cell
        tab._argH[0] = 0.5 * size_of_cell
        tab._argH[1] = 0.5 * size_of_cell

        elements.append(Paragraph("Stan techniczny nawierzchni", style))
        elements.append(tab)
        doc.build(elements)
        print("Zapisano!")

    except Exception as e:
        print(f"Błąd: {e}")
        input("Wciśnij Enter, aby zamknąć...")

window = tk.Tk()
window.title("PJP_Project")

window.rowconfigure(0, minsize=500, weight=1)
window.columnconfigure(1, minsize=500, weight=1)

frm_buttons = tk.Frame(window, relief=tk.RAISED, bd=2)
btn_open = tk.Button(frm_buttons, text="Open", command=open_file, width=30, height=2)
btn_save = tk.Button(frm_buttons, text="Save as PDF", command=save_file, width=30, height=2)

btn_open.grid(row=0, column=0, sticky="ew")
btn_save.grid(row=1, column=0, sticky="ew")

frm_buttons.grid(row=0, column=0, sticky="ns")
frm_buttons.pack()

window.mainloop()